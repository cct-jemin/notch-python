from fastapi import HTTPException
from app.config import massupload_config,v2ScopeCategoryConfig,v2ReportFieldsConfig
import logging
import pandas as pd
from app.schemas import massupload_validation_schema
import re
import asyncio
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime,timedelta
from openpyxl import load_workbook
from typing import Dict, List, Any
from collections import Counter
import copy


labelIgnore = ['category','custom']
sectionVal = {
    "section_1": "Company Information",
    "section_2": "Employee Commuting",
    "section_3": "Business Travel",
    "section_4": "Energy",
    "section_5": "Fuels",
    "section_6": "Detailed Carbon",
}
monthArr = ['January','February','March','April','May','June','July','August','September','October','November','December']
monthYearHeader = []
carElectricLabel = "car - electric",
CAR_ELECTRIC_OBJECT = {
      "kwh": {
          "fuelType": "electricity",
          "fuelSubType": "grid",
          "unit":"kwh",
      },
      "kwh - green (grid)": {
          "fuelType": "greenElectricity",
          "fuelSubType": "grid",
          "unit":"kwh - Green (Grid)",
      },
      "kwh - green (generated)": {
          "fuelType": "greenElectricity",
          "fuelSubType": "nonGrid",
          "unit":"kwh - Green (Generated)",
      }
},
PURCHASE_ENERGY_LABEL = "purchasedEnergy",
GENERATED_ENERGY_LABEL = 'generatedElectricity',
CAR_ELECTRIC_GENERATED_ELECTRICITY_UNIT  = 'kwh - green (generated)'

executor = ThreadPoolExecutor()

async def sheetWiseValidation(requestParam):
    filePath = requestParam['filePath']
    workbook = load_workbook(filePath, read_only=True)
    sheetNamesAarry = workbook.sheetnames
    sheetWiseData = {'sheets' : {}, 'validationStr' : ''}
    start_time = datetime.now()
    print(f"Start Time: {start_time.isoformat()}")
    tasks = []
    for sheetName in sheetNamesAarry:
        
        if sheetName.lower() not in massupload_validation_schema.massUploadValidationSchema["sheetMapping"] :
            continue     
    
        sheet = workbook[sheetName]
        sheetData = [[cell.value if cell.value is not None else '' for cell in row] for row in sheet.iter_rows()]
        # print(sheetData,"---")
        # tasks.append(validate_sheet(requestParam,sheetName, sheetData,sheetWiseData))
        sync_validate(requestParam,sheetName, sheetData,sheetWiseData)
        
    # Execute all tasks concurrently
    await asyncio.gather(*tasks)
    end_time = datetime.now()
    duration = (end_time - start_time).total_seconds()
    print(f"End Time: {end_time.isoformat()}")
    print(f"Duration: {duration:.2f} seconds")
        
    validation_obj = []
    if not sheetWiseData.get("sheets", {}):  # Check if the "sheets" dictionary is empty
        validation_obj.append({
            "sheetName": "No Sheets",
            "validation": ["Please upload valid sheet"]
        })
        
    for sheet_name, sheet_data in sheetWiseData.get("sheets", {}).items():
        validation_str = sheet_data.get("validationStr", "").strip()
        if validation_str:
            # Extract content between <li> and </li> tags
            matches = re.findall(r"<li>(.*?)<\/li>", validation_str)
            # Remove any extra whitespace and filter out empty values
            html_array = [match.strip() for match in matches if match.strip()]
            validation_obj.append({"sheetName": sheet_name, "validation": html_array})
            
    if validation_obj :
        return {"isAllSheetValid":False,"validationObj": validation_obj}
    else :
        if requestParam['isBenchmark']:
            # print(sheetWiseData,"aaaaaaaaaa")
            # print(list(requestParam['monthYearHeader'].keys()))
            sectionHeader = list(requestParam['monthYearHeader'].keys())
            saveSheetWiseCarbonData(requestParam, sheetWiseData, requestParam['monthYearHeader'][sectionHeader[0]])
        else :
            saveSheetWiseCarbonData(requestParam, sheetWiseData,[])
        return {"isAllSheetValid":True,"validationObj": validation_obj}


async def validate_sheet(requestParam,sheetName, sheetData,sheetWiseData):
    await asyncio.to_thread(sync_validate,requestParam,sheetName,sheetData,sheetWiseData)
   
    
    
def sync_validate(requestParam,sheetName, sheetData, sheetWiseData):
    # start_time = time.time()
    # print(f"Start processing sheet: {sheetName} at {start_time:.2f}")
    sectionWiseSheetData = {}
    maxAllowedColumnNumber = 25
    headerStartIndexColumnNumber = 13
    headerPeriodArr = []
    requestParam.update({
        'maxAllowedColumnNumber': maxAllowedColumnNumber,
        'headerStartIndexColumnNumber': headerStartIndexColumnNumber,
        'sheetName': sheetName,
        'headerPeriodArr':headerPeriodArr
    })
    sheetInfo = massupload_validation_schema.massUploadValidationSchema["sheetMapping"][sheetName.lower()]
    sectionNameMapping = massupload_validation_schema.massUploadValidationSchema['sectionNameMapping']
    sectionNameMappingReverse = reverseObject(sectionNameMapping)
   #for now take data from config file
    scope = str(sheetInfo['scope'])
    category = sheetInfo['category']
    requestParam.update({'scope': scope, 'category': category})
    sheetWiseData['sheets'][sheetName] = {"scope" : scope, "category" : category, "data" : [], "sectionPointer" : '', "validationStr" : ''}
    section_pointer_name = None
    section_pointer = ''
    valid_section_name_array = [
        sectionNameMappingReverse[sectionPointer]
        for sectionPointer in sheetInfo['sectionPointer']
        if sectionPointer in sectionNameMappingReverse
    ]
    headerRow = []
    for i, row in enumerate(sheetData):
        # validation of row 1 column 1.
        if i == 0:
            first_cell_value = lower_case(row[0])
            if not category and first_cell_value != "company information":
                sheetWiseData['sheets'][sheetName]['validationStr'] += (
                    f"<li>Row - 1 : Invalid names. It should be Company Information.</li>"
                )
                break    
            elif not category and first_cell_value == "company information":
                section_pointer_name = first_cell_value
            elif category and first_cell_value == "category":
                section_pointer_name = lower_case(sheetData[1][0]) if len(sheetData) > 1 else None
            else:
                sheetWiseData['sheets'][sheetName]['validationStr'] += (
                    f"<li>Row - 1 : Invalid {row[0]}. It should be Category.</li>"
                )
                break  
            
        # Filter out rows where all cells are null
        if all(cell is None or cell == '' for cell in row):
            continue

        # Trim strings in the row
        row = [cell.strip() if isinstance(cell, str) else cell for cell in row]
        
        # Slice columns to only include the maximum allowed number
        row = row[:maxAllowedColumnNumber]
        sheetData[i] = row
        
        # section validation
        if row[0] and row[0].lower() in valid_section_name_array:
            section_pointer_name = row[0].lower()
            section_pointer = sectionNameMapping[section_pointer_name]
            sheetWiseData['sheets'][sheetName]['sectionPointer'] = section_pointer
            # requestParam['sectionPointer'] = sectionNameMapping[section_pointer_name]
            requestParam['sectionPointer'] = section_pointer

        elif row[0] and section_pointer_name != 'company information' and row[0].lower() not in labelIgnore:
            sheetWiseData['sheets'][sheetName]['validationStr'] += (
                f"<li>Row - {i+1} : Invalid Section name. It should be one of these {', '.join(valid_section_name_array)}.</li>"
            )
            break
        
        # Category validation
        if section_pointer_name != 'company information':
            attributeCategoryFields = v2ScopeCategoryConfig.scopeConfig[scope][category]
            attributeSubCategoryFields = get_attribute_subcategories(v2ScopeCategoryConfig.scopeConfig)
            requestParam['attributeCategoryFields'] = attributeCategoryFields
            # print(lower_case(sheetData[i][0]),"aaa")
            # print(section_pointer_name,"section_pointer_name")
            if row[0] and row[0].strip().lower() == "category":
                # Check if the attributeCategoryFields label matches the category in sheet_data
                if attributeCategoryFields and attributeCategoryFields['label'].strip().lower() != sheetData[0][1].strip().lower():
                    sheetWiseData['sheets'][sheetName]['validationStr'] += f"<li>Row - {i + 1} : Invalid Category name {sheetData[0][1]}. It should be {attributeCategoryFields['label']}.</li>"
                    break
            elif  lower_case(sheetData[i][0]) == 'custom' or lower_case(sheetData[i][2]) == 'custom' :
                if attributeSubCategoryFields and sheetData[i][1] in attributeSubCategoryFields :
                    sheetWiseData['sheets'][sheetName]['validationStr'] =  f"<li>Row - {i + 1} : Custom project sub-category name should not be same as standard sub-category {sheetData[i][1]}.</li>"
                    break
                else :
                    sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseSectionPropertiesValidation(requestParam, section_pointer, sheetData, i)
                    sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseValueValidation(requestParam, sheetData, i, headerRow)
                
            elif  lower_case(sheetData[i][0]) == section_pointer_name :
                headers = massupload_validation_schema.massUploadValidationSchema['headerMapping'][section_pointer]
                if headers:
                    headerRow = sheetData[i]
                    sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseHeaderValidation(requestParam,section_pointer, headers, sheetData, i)
                    sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseHeaderPeriodValidation(requestParam, section_pointer, sheetData, i)
               
            else :
                #Property and value validation
                sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseSectionPropertiesValidation(requestParam, section_pointer, sheetData, i)
                sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseValueValidation(requestParam, sheetData, i,headerRow)
        
        elif  lower_case(sheetData[i][0]) == section_pointer_name :
                headers = massupload_validation_schema.massUploadValidationSchema['headerMapping'][section_pointer]
                if headers:
                    headerRow = sheetData[i]
                    #Header validation
                    sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseHeaderValidation(requestParam,section_pointer, headers, sheetData, i)
                    sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseHeaderPeriodValidation(requestParam, section_pointer, sheetData, i)
        else :
            #Property and value validation
            sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseSectionPropertiesValidation(requestParam, section_pointer, sheetData, i)
            sheetWiseData['sheets'][sheetName]['validationStr'] += sheetWiseValueValidation(requestParam, sheetData, i,headerRow)
            
        if section_pointer not in sectionWiseSheetData:
            sectionWiseSheetData[section_pointer] = []

        sectionWiseSheetData[section_pointer].append(sheetData[i])
            
    sheetWiseData['sheets'][sheetName]['data'] = sectionWiseSheetData;   
    if not requestParam['isBenchmark']:
        if len(requestParam['monthYearHeader']) == 12:
            monthYearDates = sorted(
                datetime.strptime(value + '-01', '%b-%Y-%d') for value in requestParam['monthYearHeader']
            )
            yearMonth = '202101'
            benchmarkDate = datetime.strptime(yearMonth[:4] + '-' + yearMonth[4:] + '-01', '%Y-%m-%d')
            if benchmarkDate > monthYearDates[0]:
                sheetWiseData['sheets'][sheetName]['validationStr'] += f"<li>Data should be greater or equal to your site benchmark period.</li>"
            elif benchmarkDate < monthYearDates[0]:
                previousMonth = (monthYearDates[0] - timedelta(days=1)).strftime('%b-%Y')
                checkExists = True
                if not checkExists:
                    sheetWiseData['sheets'][sheetName]['validationStr'] +=  f'<li>Please insert {previousMonth} month-year data first.</li>'
     
    # if  requestParam['isBenchmark']:
    #     print(requestParam['headerPeriodArr'],"1111111111")  
                
                
                

    # sheetWiseData['sheets'][sheetName]['data'] = sheetData
    # print(f"Finished processing sheet: {sheetName} at {time.time():.2f}, duration: {time.time() - start_time:.2f} seconds")
    
def sheetWiseHeaderValidation(request_param,section_pointer, headers, sheetData, i):
    validation_str = ""
    actual_header = []
    
    # Validate column length
    if len(sheetData[i]) != request_param['maxAllowedColumnNumber']:
        validation_str += (
            f"<li>Row - {i + 1}: Section ({sectionVal[section_pointer]}) Header Not valid </li>"
        )

    # Slice header based on section
    if section_pointer == "section_1":
        actual_header = sheetData[i][:2]
    elif section_pointer in {"section_2", "section_3", "section_4", "section_5"}:
        actual_header = sheetData[i][:9]
    elif section_pointer == "section_6":
        actual_header = sheetData[i][:7]
    

    # Normalize header to lowercase and remove None values
    actual_header = [
        str(name).lower() for name in actual_header if name is not None
    ]
    
    # Find the difference between valid and actual headers
    diff_header = list(set(headers) - set(actual_header))
    if diff_header:
        diff_header = [header.capitalize() for header in diff_header]
        validation_str += (
            f"<li>Row - {i + 1}: Section ({sectionVal[section_pointer]}) ["
            + ", ".join(diff_header)
            + "] Header Not found </li>"
        )

    return validation_str

def sheetWiseHeaderPeriodValidation(requestParam, sectionPointer, sheetData, rowIndex):
    validationStr = ''
    periodHeaderArr = []

    def convertAndValidateHeader(header):
        if header == '':
            return None
        try:
            if isinstance(header, (int, float)):
                headerDate = excelDateToFormatedDate(header).strftime("%b-%Y").lower()
            elif isinstance(header, str):
                headerDate = header.strip().lower()
                if not validateDate(headerDate):
                    raise ValueError(f"{header}: Header Not Valid. Follow this format: (Jan-2019)")
            else:
                raise ValueError("Header cannot be None")
            return headerDate
        except ValueError as e:
            nonlocal validationStr
            validationStr += f"<li>Row - {rowIndex + 1}: {str(e)}</li>"
            return None

    # Validate and process header periods
    for header in sheetData[rowIndex][requestParam['headerStartIndexColumnNumber']:]:
        headerPeriod = convertAndValidateHeader(header)
        if headerPeriod:  # Only add if the header is valid
            periodHeaderArr.append(headerPeriod)
            
    if len(periodHeaderArr) < 12:
        validationStr += f"<li>Row - {rowIndex + 1}: Must have at least 12 months and year headers.</li>"
        

    # Validate sorted and current date constraints
    sortedDates = sorted(
        [datetime.strptime(h, "%b-%Y") for h in periodHeaderArr if h],
        reverse=False
    )
    if sortedDates and sortedDates[-1] > datetime.now():
        validationStr += f"<li>Row - {rowIndex + 1}: Month and year headers must be prior to the current month/year.</li>"

    # Check for duplicate headers
    if periodHeaderArr:
        monthYearHeader = [re.sub(r'^b-', '', h) for h in periodHeaderArr if h]
        duplicates = [item for item, count in Counter(periodHeaderArr).items() if count > 1]
        if duplicates:
            validationStr += f"<li>Row - {rowIndex + 1}: Duplicate headers found ({', '.join(duplicates)}).</li>"
            
        # print(periodHeaderArr,"periodHeaderArr")
        monthYearDates = sorted(
            datetime.strptime(value + '-01', '%b-%Y-%d') for value in periodHeaderArr
        )
    
        yearMonth = '202101'
        benchmarkDate = datetime.strptime(yearMonth[:4] + '-' + yearMonth[4:] + '-01', '%Y-%m-%d')
            
        # if benchmarkDate > monthYearDates[0]:
        #     validationStr += f"<li>Row - {rowIndex + 1}: Benchmark Period should be greater or equal to your site benchmark period.</li>"

        if benchmarkDate == monthYearDates[0]:
            requestParam['isBenchmark'] = True 
            requestParam['headerPeriodArr'] =  list(set(requestParam['headerPeriodArr'] + monthYearHeader))
            if not requestParam.get('monthYearHeader'):
                requestParam['monthYearHeader'] = {}
                
            if sectionPointer not in requestParam['monthYearHeader']:
                requestParam['monthYearHeader'][sectionPointer] = []
                
            requestParam['monthYearHeader'][sectionPointer] = periodHeaderArr
            
        else :
            requestParam['isBenchmark'] = False 
             # Match headers with global monthYearHeader
            if not requestParam.get('monthYearHeader'):
                requestParam['monthYearHeader'] =  [re.sub(r'^b-', '', h) for h in periodHeaderArr if h]
                requestParam['baseHeader'] = sectionPointer
            else:
                # currentHeaders = [h.replace("b-", "") for h in periodHeaderArr if h]
                currentHeaders = [re.sub(r'^b-', '', h) for h in periodHeaderArr if h]
                if not set(requestParam['monthYearHeader']) == set(currentHeaders):
                    validationStr += (
                        f"<li>Row - {rowIndex + 1}: Month-year headers should match with "
                        f"{requestParam['baseHeader']} month-year headers.</li>"
                    )
   

    return validationStr

def sheetWiseValueValidation(requestParam, sheetData, rowIndex,headerRow):
    validationStr = ''
    isValidate = True
    for index in range(requestParam['headerStartIndexColumnNumber'], len(sheetData[rowIndex])):
        columnName = headerRow[index]
        if isinstance(headerRow[index], (int, float)):
            columnName = excelDateToFormatedDate(headerRow[index]).strftime("%b-%Y")

        cellValue = sheetData[rowIndex][index]
        if isValidate:
            if cellValue == '':
                validationStr += f"<li>Row - {rowIndex + 1} : Value required in {columnName} column.</li>"
            # Check if the value is less than 0
            elif isinstance(cellValue, (int, float)) and cellValue < 0:
                validationStr += f"<li>Row - {rowIndex + 1} : Value must be 0 or higher in {columnName} column.</li>"
            

        # Check if the value is not numeric
        if cellValue != '' and not isinstance(cellValue, (int, float)):
            validationStr += f"<li>Row - {rowIndex + 1} : Value must be numeric in {columnName} column.</li>"

    return validationStr

def sheetWiseSectionPropertiesValidation(requestParam, sectionPointer, sheetData, rowIndex):
    validationStr = ''
    if sectionPointer == 'section_1':
        companyInfoField = sheetData[rowIndex][0]
        companyInfoCustomField = sheetData[rowIndex][1]
        sectionFields = v2ReportFieldsConfig.fieldsConfig['companyInfoFields']
        
        sectionLabel = ['custom 1', 'custom 2', 'custom 3'] + [
            field['label'].strip().lower() for field in sectionFields.values()
        ]
        # Convert the companyInfoField to lowercase for case-insensitive comparisons
        if companyInfoField is not None:
            companyInfoFieldStr = companyInfoField.strip().lower()

            # Validate if companyInfoField matches any valid label
            if companyInfoFieldStr not in sectionLabel:
                validationStr += (
                    f"<li>Row - {rowIndex + 1}: ({sectionVal[sectionPointer]}) "
                    f"{companyInfoField.strip()} value not match with {', '.join(sectionLabel)}.</li>"
                )

            # Custom field-specific validations
            if companyInfoFieldStr.startswith('custom'):
                if not companyInfoCustomField:
                    validationStr += (
                        f"<li>Row - {rowIndex + 1}: ({sectionVal[sectionPointer]}) "
                        f"custom type must be required for {companyInfoField}.</li>"
                    )
            elif companyInfoCustomField:
                validationStr += (
                    f"<li>Row - {rowIndex + 1}: ({sectionVal[sectionPointer]}) "
                    f"custom type should be blank for {companyInfoField}.</li>"
                )
    elif sectionPointer == 'section_2' or sectionPointer == 'section_3' :
        SheetDataObj = {
            "attributeSubCategory": sheetData[rowIndex][1] if sheetData[rowIndex][1] else "blank",
            "type": sheetData[rowIndex][2],
            "subType": sheetData[rowIndex][3],
            "unit": sheetData[rowIndex][4] if sheetData[rowIndex][4] else "blank",
        }
        validationStr +=  validateFields(requestParam , SheetDataObj, sheetData, rowIndex)
        params = [
            {
            'fieldName' : 'Cost/Unit',
            'required' : True,
            'type' : 'number',
            'sheetData' : sheetData[rowIndex][5]
            },
            {
            'fieldName' : 'Driver Name',
            'required' : False,
            'type' : 'string',
            'sheetData' : sheetData[rowIndex][6]
            },
            {
            'fieldName' : 'Registration',
            'required' : False,
            'type' : 'string',
            'sheetData' : sheetData[rowIndex][7]
            },
        ]
        validationStr +=  fieldTypeValidation(params,rowIndex)
    elif sectionPointer == 'section_4' :
        energyFields = ["green electricity", "electricity", "gas" ,"steam"]
        SheetDataObj = {
            'attributeSubCategory' : sheetData[rowIndex][1] if sheetData[rowIndex][1] else "blank",
            'type' : sheetData[rowIndex][2],
            'subType' : sheetData[rowIndex][3],
            'unit' : sheetData[rowIndex][4] if sheetData[rowIndex][4] else "blank",
            'workplace' : sheetData[rowIndex][6] if sheetData[rowIndex][6] else "blank",
            'meterName' : sheetData[rowIndex][7],
            'meterNumber' : sheetData[rowIndex][8]
        }
        if SheetDataObj['attributeSubCategory'].strip().lower() not in energyFields:
            validationStr += (
                f"<li>Row - {rowIndex + 1}: Invalid Sub-Category ({SheetDataObj['attributeSubCategory']}) "
                f"It should be one of these {energyFields}.</li>"
            )
        else :
            validationStr +=  validateFields(requestParam , SheetDataObj, sheetData, rowIndex)
            params = [
                {
                'fieldName' : 'Cost/Unit',
                'required' : True,
                'type' : 'number',
                'sheetData' : sheetData[rowIndex][5]
                },
                {
                'fieldName' : 'Meter Name',
                'required' : False,
                'type' : 'string',
                'sheetData' : sheetData[rowIndex][7]
                },
                {
                'fieldName' : 'Meter Number',
                'required' : False,
                'type' : 'string',
                'sheetData' : sheetData[rowIndex][8]
                },
            ]
            validationStr +=  fieldTypeValidation(params, rowIndex)
    elif sectionPointer == 'section_5' :
        SheetDataObj = {
            'attributeSubCategory' : sheetData[rowIndex][1] if sheetData[rowIndex][1] else "blank",
            'type' : sheetData[rowIndex][2] if sheetData[rowIndex][1] else "blank",
            'subType' : sheetData[rowIndex][3],
            'unit' : sheetData[rowIndex][4] if sheetData[rowIndex][4] else "blank",
            'workplace' : sheetData[rowIndex][6] if sheetData[rowIndex][6] else "blank",
            'customName' : sheetData[rowIndex][7],
            'carbonFactor' : sheetData[rowIndex][8],
        }
        fuelsFields  = ['fuels', 'fugitive emissions']
        if SheetDataObj['attributeSubCategory'].strip().lower() not in fuelsFields:
            validationStr += (
                f"<li>Row - {rowIndex + 1}: Invalid Sub-Category ({SheetDataObj['attributeSubCategory']}) "
                f"It should be one of these {fuelsFields}.</li>"
            )
        else : 
            validationStr +=  validateFields(requestParam , SheetDataObj, sheetData, rowIndex)
            params = [
                {
                    'fieldName' : 'Cost/Unit',
                    'required' : True,
                    'type' : 'number',
                    'sheetData' : sheetData[rowIndex][5]
                }
            ]
            validationStr +=  fieldTypeValidation(params, rowIndex)
    elif sectionPointer == 'section_6' :
        SheetDataObj = {
            'attributeType': sheetData[rowIndex][0],
            'attributeSubCategory' : sheetData[rowIndex][1],
            'type' : sheetData[rowIndex][2],
            'subType' : sheetData[rowIndex][3] ,
            'unit' : sheetData[rowIndex][4],
        }
        if isinstance(SheetDataObj['type'], str) and SheetDataObj['type'].strip().lower() == 'custom' or SheetDataObj['attributeType'].strip().lower() == 'custom' :
            isTypeRequired = True
            if SheetDataObj['attributeType'].strip().lower() == 'custom' and (not SheetDataObj.get('subType') or SheetDataObj['subType'] is None):
                isTypeRequired = False
                
            params = [
                {
                'fieldName' : 'Sub - Category',
                'required' : True,
                'type' : 'string',
                'sheetData' : SheetDataObj['attributeSubCategory']
                },
                {
                'fieldName' : 'Type',
                'required' : isTypeRequired,
                'type' : 'string',
                'sheetData' : SheetDataObj['type']
                },
                {
                'fieldName' : 'sub-type',
                'required' : False,
                'type' : 'string',
                'sheetData' : SheetDataObj['subType']
                },
                {
                'fieldName' : 'Unit',
                'required' : True,
                'type' : 'string',
                'sheetData' : SheetDataObj['unit']
                },
                {
                'fieldName' : 'Co2 Unit/Kg',
                'required' : True,
                'type' : 'number',
                'sheetData' : sheetData[rowIndex][6]
                }
            ]
            validationStr +=  fieldTypeValidation(params,rowIndex)
        else :
            validationStr +=  validateFields(requestParam , SheetDataObj, sheetData, rowIndex)
            params = [
                {
                    'fieldName' : 'Co2 Unit/Kg',
                    'required' : '',
                    'type' : 'number',
                    'sheetData' : sheetData[rowIndex][6]
                },
            ]
            validationStr +=  fieldTypeValidation(params, rowIndex)
        params = [
            {
            'fieldName' : 'Cost/Unit',
            'required' : True,
            'type' : 'number',
            'sheetData' : sheetData[rowIndex][5]
            }
        ]
        validationStr +=  fieldTypeValidation(params, rowIndex)

    return validationStr


def validateFields(request_param, sheet_data_obj, sheet_data, row_index):
    section_fields = request_param["attributeCategoryFields"].get("attributeSubCategory", {})
    validation_str = ''
    attribute_sub_category_valid = False
    is_type_valid = False
    is_sub_type_valid = False
    attribute_sub_category_labels = []
    custom_project_found = False
    category = request_param["category"]
    scopeValue = request_param.get("scope")

    # Attribute sub-category validation
    if section_fields != "userdefine":
        # workplace validation
        if request_param["sectionPointer"] in ["section_4", "section_5"]:
            workplaceSheet = sheet_data_obj['workplace'].strip().lower()
            if workplaceSheet not in ['home', 'office']:
                validation_str += (
                    f"<li>Row - {row_index + 1}: Invalid workplace {sheet_data_obj['workplace']}. "
                    f"It should be home or office.</li>"
                )
            elif scopeValue == "1" and workplaceSheet != "office":
                validation_str += (
                    f"<li>Row - {row_index + 1}: Invalid workplace {sheet_data_obj['workplace']}. "
                    f"It should be office.</li>"
                )
            elif scopeValue == "2" and sheet_data_obj["attributeSubCategory"].strip().lower() in ["fuels", "fugitive emissions"] and workplaceSheet != "home":
                validation_str += (
                    f"<li>Row - {row_index + 1}: Invalid workplace {sheet_data_obj['workplace']}. "
                    f"It should be home.</li>"
                )
            elif scopeValue == "3" and workplaceSheet != "home":
                validation_str += (
                    f"<li>Row - {row_index + 1}: Invalid workplace {sheet_data_obj['workplace']}. "
                    f"It should be home22.</li>"
                )

            # Validation for meter name, meter number, and sub-type for home workplace
            if request_param["sectionPointer"] == "section_4" and workplaceSheet == "home":
                if sheet_data_obj.get("meterName"):
                    validation_str += (
                        f"<li>Row - {row_index + 1}: Meter Name should be blank for home.</li>"
                    )
                if sheet_data_obj.get("meterNumber"):
                    validation_str += (
                        f"<li>Row - {row_index + 1}: Meter Number should be blank for home.</li>"
                    )
                if sheet_data_obj.get("subType"):
                    validation_str += (
                        f"<li>Row - {row_index + 1}: Sub-type should be blank for home.</li>"
                    )
            
        for key, section_field in section_fields.items():
            if key == "custom":
                custom_project_found = True
            
            fieldData = copy.deepcopy(section_field)
            attribute_sub_category_field = section_field.get("sheetLabel") or section_field.get("label")
            attribute_sub_category_labels.append(attribute_sub_category_field)
            if attribute_sub_category_field.strip().lower() == sheet_data_obj["attributeSubCategory"].strip().lower():
                attribute_sub_category_valid = True
                type_labels = []

                # Handling 'plane' specific logic
                if key == "plane":
                    attributeSubCategoryPlane = copy.deepcopy(fieldData['type'])
                    fieldData["type"] = {
                        "domestic": attributeSubCategoryPlane["plane"]["options"]["domestic"],
                        "shortHaul": attributeSubCategoryPlane["plane"]["options"]["shortHaul"],
                        "longHaul": attributeSubCategoryPlane["plane"]["options"]["longHaul"],
                        "international": attributeSubCategoryPlane["plane"]["options"]["international"],
                    }
                    if request_param["sectionPointer"] == "section_5":
                        fieldData["type"].update({
                            "economy": attributeSubCategoryPlane["plane"]["options"]["economy"],
                            "business": attributeSubCategoryPlane["plane"]["options"]["business"]
                        })
                    fieldData["unit"] = attributeSubCategoryPlane["plane"]["unit"]
                    
                    
                for type_key, type_field in fieldData["type"].items():
                    typeFields = copy.deepcopy(type_field)
                    # print(typeFields,"11111111111111111111111111111111111")
                    type_label = typeFields.get("sheetLabel") or typeFields.get("label")
                    type_labels.append(type_label)
                    options = typeFields.get("options", {})
                    if type_label.strip().lower() == sheet_data_obj["type"].strip().lower():
                        is_type_valid = True
                        units = [unit.strip().lower() for unit in typeFields.get("unit", [])]
                        if typeFields.get("unit", []) and len(options) == 0 and sheet_data_obj["unit"].strip().lower() not in units:
                            validation_str += (
                                f"<li>Row - {row_index + 1}: Invalid unit {sheet_data_obj['unit']}. "
                                f"It should be one of these: {', '.join(units)}.</li>"
                            )
                            
                        if sheet_data_obj["attributeSubCategory"].strip().lower() == 'plane':
                            attributeSubCategoryUnit = fieldData['unit']
                            if sheet_data_obj["unit"].strip().lower() not in attributeSubCategoryUnit:
                                validation_str += (
                                    f"<li>Row - {row_index + 1}: Invalid unit {sheet_data_obj['unit']}. "
                                    f"It should be one of these: {', '.join(attributeSubCategoryUnit)}.</li>"
                                )
                               
                        subtype_labels = []
                        for sub_type_key, sub_type_field in options.items():
                            subTypeFields = typeFields['options'][sub_type_key]
                            # print( typeFields['options'],"qqqq")
                            sub_type_label = subTypeFields.get("sheetLabel") or subTypeFields.get("label")
                            subtype_labels.append(sub_type_label)
                            if sub_type_label.strip().lower() == sheet_data_obj["subType"].strip().lower():
                                is_sub_type_valid = True
                                # print("innnnn",subTypeFields)
                                # Validate energy source
                                if category == "purchasedEnergy" and "generated" in subTypeFields.get("source", []):
                                    validation_str += (
                                        f"<li>Row - {row_index + 1}: Invalid subtype {sheet_data_obj['subType']}. "
                                        f"It should be {options['grid']['label']}.</li>"
                                    )
                                elif category == "generatedElectricity" and "non-generated" in subTypeFields.get("source", []):
                                    generated_labels = [
                                        option["label"]
                                        for option in options.values()
                                        if "generated" in option.get("source", [])
                                    ]
                                    validation_str += (
                                        f"<li>Row - {row_index + 1}: Invalid subtype {sheet_data_obj['subType']}. "
                                        f"It should be one of these: {', '.join(generated_labels)}.</li>"
                                    )
                                # Validate unit for sub-type
                                if subTypeFields.get("unit", []) :
                                    units = [unit.strip().lower() for unit in subTypeFields.get("unit", [])]
                                else :
                                    units = [unit.strip().lower() for unit in typeFields.get("unit", [])]
                                    
                                if sheet_data_obj["unit"].strip().lower() not in units:
                                    validation_str += (
                                        f"<li>Row - {row_index + 1}: Invalid unit {sheet_data_obj['unit']}. "
                                        f"It should be one of these: {', '.join(units)}.</li>"
                                    )

                        if not is_sub_type_valid and subtype_labels:
                            validation_str += (
                                f"<li>Row - {row_index + 1}: Invalid sub-type {sheet_data_obj['subType'] or 'blank'}. "
                                f"It should be one of these: {', '.join(subtype_labels)}.</li>"
                            )

                if not is_type_valid and type_labels:
                    validation_str += (
                        f"<li>Row - {row_index + 1}: Invalid type {sheet_data_obj['type'] or 'blank'}. "
                        f"It should be one of these: {', '.join(type_labels)}.</li>"
                    )

        if not attribute_sub_category_valid:
            validation_str += (
                f"<li>Row - {row_index + 1}: Invalid sub-category {sheet_data_obj['attributeSubCategory']}. "
                f"It should be one of these: {', '.join(attribute_sub_category_labels)}.</li>"
            )
            if custom_project_found:
                validation_str += (
                    f"<li>Row - {row_index + 1}: For Custom project, 'Custom' should be added in 'Detailed Carbon' column.</li>"
                )

    elif section_fields == "userdefine":
        validation_str += (
            f"<li>Row - {row_index + 1}: Invalid type {sheet_data_obj['type']}. It should be 'Custom'.</li>"
        )
    else:
        validation_str += (
            f"<li>Row - {row_index + 1}: Invalid sub-category {sheet_data_obj['attributeSubCategory']}.</li>"
        )

    return validation_str

def fieldTypeValidation(params,rowIndex):
    validationStr = ""
    if params:
        for obj in params:
            fieldName = obj.get('fieldName', 'unknown')
            sheetDataValue = obj.get('sheetData')
            required = obj.get('required', False)
            expectedType = obj.get('type')

            # Convert expected type to Python type
            pythonType = str if expectedType == 'string' else (int, float) if expectedType == 'number' else None
            # Check if sheetDataValue matches expected type
            if pythonType and sheetDataValue and not isinstance(sheetDataValue, pythonType):
                validationStr += (
                    f"<li>Row - {rowIndex + 1}: Invalid field {fieldName}. "
                    f"It shoulds be {expectedType}.</li>"
                )

            # Handle required validations
            if required:
                if expectedType == 'number' and not isinstance(sheetDataValue, (int, float)):
                    validationStr += (
                        f"<li>Row - {rowIndex + 1}: Invalid field {fieldName}. "
                        f"It should not be blank and should be a number.</li>"
                    )
                elif expectedType == 'string':
                    if isinstance(sheetDataValue, (int, float)):
                        validationStr += (
                            f"<li>Row - {rowIndex + 1}: Invalid field {fieldName}. "
                            f"It should not be a number.</li>"
                        )
                    elif not sheetDataValue:
                        validationStr += (
                            f"<li>Row - {rowIndex + 1}: Invalid field {fieldName}. "
                            f"It should not be blank.</li>"
                        )
            # Handle non-required fields with unexpected data
            elif required == '' and sheetDataValue:
                validationStr += (
                    f"<li>Row - {rowIndex + 1}: Invalid field {fieldName}. "
                    f"It shouldss be blank.</li>"
                )

    return validationStr
def saveSheetWiseCarbonData(request_param, sheet_wise_data, month_year_header=[]):
    final_obj = {}
    delete_custom = {
        'custom_1': True,
        'custom_2': True,
        'custom_3': True
    }
    
    benchmark_period = [
        '202101', '202102', '202103', '202104', '202105', '202106',
        '202107', '202108', '202109', '202110', '202111', '202112'
    ]
    
    if request_param.get('isBenchmark') and len(month_year_header) == 12:
        request_param['monthYearHeader'] = month_year_header
        
    for yearcnt in range(len(request_param['monthYearHeader'])):
        section_one = {}
        section_two = {}
        section_three = {}
        section_four = {
            'electricity': [],
            'gas': []
        }
        section_five = {}
        month_str = request_param['monthYearHeader'][yearcnt][:3]
        month_no = get_month(month_str)
        year_month = int(request_param['monthYearHeader'][yearcnt].split('-')[1])
        year_month = f'{year_month}{month_no}'
        direct_carbon = None
        
        if year_month in benchmark_period:
            request_param['isBenchmark'] = True
        
        if month_no is not None:
            final_obj['monthName'] = monthArr[month_no]
            final_obj['month'] = month_no + 1
            month = f"{final_obj['month']:02d}"  # Zero-padded month (e.g., 01 for Jan)
            final_obj['year'] = int(request_param['monthYearHeader'][yearcnt].split('-')[1])
            final_obj['yearMonth'] = f"{final_obj['year']}{month}"
            final_obj['total_carbon'] = None
        
        # print(sheet_wise_data['sheets'].items(),"111111111")
        for sheet_name, sheet_data in sheet_wise_data['sheets'].items():
            category_row = ''
            section_wise_data = {}
            for data_row,pointer in enumerate(sheet_data['data']):
                if sheet_data['data'][pointer][0][0].lower() == 'category':
                    category_row = sheet_data['data'][pointer][0]
                elif category_row:
                    sheet_data['data'][pointer].insert(0, category_row)

                section_wise_data[pointer] = sheet_data['data'][pointer]
                category = sheet_data['category']
                scope = sheet_data['scope']

                # Section 1 (Company Info)
                if 'section_1' in section_wise_data:
                    for s_cnt1 in range(1, len(section_wise_data['section_1'])):
                        row_data = section_wise_data['section_1'][s_cnt1]
                        company_info_row = row_data[0]
                        custom_type_row = row_data[1]
                        value = row_data[yearcnt + request_param['headerStartIndexColumnNumber']]

                        if company_info_row.lower().startswith('custom'):
                            if 'custom' not in section_one:
                                section_one['custom'] = {}

                            if company_info_row.lower() == 'custom 1':
                                section_one['custom']['custom_1'] = {
                                    'metric': custom_type_row,
                                    'value': value,
                                    'icon': ""
                                }
                                delete_custom['custom_1'] = custom_type_row is None

                            if company_info_row.lower() == 'custom 2':
                                section_one['custom']['custom_2'] = {
                                    'metric': custom_type_row,
                                    'value': value,
                                    'icon': ""
                                }
                                delete_custom['custom_2'] = custom_type_row is None

                            if company_info_row.lower() == 'custom 3':
                                section_one['custom']['custom_3'] = {
                                    'metric': custom_type_row,
                                    'value': value,
                                    'icon': ""
                                }
                                delete_custom['custom_3'] = custom_type_row is None
                        else:
                            car_type = re.sub(r' - ', '', f"comp {company_info_row}").replace(' ', '_').lower()
                            section_one[car_type] = value

                    if section_wise_data['section_1'] and section_one:
                        section_one['is_benchmark'] = request_param.get('isBenchmark', False)
                        
                # Employee Commuting
                if 'section_2' in section_wise_data:
                    for s_cnt1 in range(2, len(section_wise_data['section_2'])):
                        row_data = section_wise_data['section_2'][s_cnt1]
                        sub_category = row_data[1]
                        type_ = row_data[2]
                        sub_type = row_data[3]
                        unit = row_data[4]
                        cost_per_unit = row_data[5]
                        driver_name = row_data[6]
                        registration_number = row_data[7]
                        is_reporting = row_data[8]
                        value = row_data[yearcnt + request_param['headerStartIndexColumnNumber']]

                        car_type = re.sub(r' - ', '', re.sub(' ', '_', type_.lower()))
                        travel_vehicle = re.sub(' ', '_', sub_type.lower())
                        travel_type = ""

                        if car_type.startswith('car'):
                            travel_vehicle = f"car_{travel_vehicle}"

                        if sub_category.lower() == 'plane':
                            car_type = sub_category.lower()
                            travel_type = re.sub(' ', '_', type_.lower())
                            travel_vehicle = re.sub(' ', '_', sub_type.lower())

                        if unit and unit not in [None, '']:
                            unit = unit.lower()

                        if car_type not in section_two:
                            section_two[car_type] = []

                        inner_object = {
                            'attributeCategory': category,
                            'attributeSubCategory': re.sub(' ', '_', sub_category.lower()),
                            'travelMode': car_type,
                            'travelVehicle': travel_vehicle,
                            'travelType': travel_type,
                            'travelValue': value,
                            'travel_id': None,
                            'name': driver_name,
                            'registration_number': registration_number,
                            'costPerUnit': cost_per_unit,
                            'scope': scope,
                            'unit': unit,
                            'isBenchmark': request_param.get('isBenchmark', False),
                            'source': 'massUpload',
                            'modifiedDate': datetime.now()
                        }

                        if type_.strip().lower() == carElectricLabel:
                            electric_data = CAR_ELECTRIC_OBJECT.get(unit, {})
                            if electric_data:
                                inner_object.update({
                                    'fuelType': electric_data.get('fuelType'),
                                    'fuelSubType': electric_data.get('fuelSubType'),
                                    'unit': electric_data.get('unit')
                                })

                        if is_reporting and is_reporting.lower() == 'yes':
                            section_two[car_type].append(inner_object)

                    if section_wise_data['section_2'] and section_two:
                        section_two['is_benchmark'] = request_param.get('isBenchmark', False)
                        
                # Business Travel 
                if 'section_3' in section_wise_data:
                    for s_cnt1 in range(2, len(section_wise_data['section_3'])):
                        row_data = section_wise_data['section_3'][s_cnt1]
                        sub_category = row_data[1]
                        type_ = row_data[2]
                        sub_type = row_data[3]
                        unit = row_data[4]
                        cost_per_unit = row_data[5]
                        driver_name = row_data[6]
                        registration_number = row_data[7]
                        is_reporting = row_data[8]
                        value = row_data[yearcnt + request_param['headerStartIndexColumnNumber']]
                        travel_vehicle = ''
                        travel_type = ''
                        
                        car_type = re.sub(r' - ', '', re.sub(' ', '_', type_.lower()))
                        travel_vehicle = re.sub(' ', '_', sub_type.lower())
                        if car_type.startswith('car'):
                            travel_vehicle = f"car_{travel_vehicle}"

                        if sub_category.lower() == 'plane':
                            car_type = sub_category.lower()
                            travel_type = re.sub(' ', '_', type_.lower())
                            travel_vehicle = re.sub(' ', '_', sub_type.lower())

                        if unit and unit not in [None, '']:
                            unit = unit.lower()

                        if car_type not in section_three:
                            section_three[car_type] = []
                         
                        inner_object = {
                            'attributeCategory': category,
                            'attributeSubCategory': re.sub(' ', '_', sub_category.lower()),
                            'travelMode': car_type,
                            'travelVehicle': travel_vehicle,
                            'travelType': travel_type,
                            'travelValue': value,
                            'travel_id': None,
                            'name': driver_name,
                            'registration_number': registration_number,
                            'costPerUnit': cost_per_unit,
                            'scope': scope,
                            'unit': unit,
                            'isBenchmark': request_param.get('isBenchmark', False),
                            'source': 'massUpload',
                            'modifiedDate': datetime.now()
                        }
                        
                        if type_.strip().lower() == carElectricLabel:
                            if scope == 1 :
                                inner_object.update({'scope':2,'attributeCategory' : PURCHASE_ENERGY_LABEL})
                                if unit == CAR_ELECTRIC_GENERATED_ELECTRICITY_UNIT :
                                    inner_object.update({'attributeCategory' : GENERATED_ENERGY_LABEL})
                                   
                            
                            electric_data = CAR_ELECTRIC_OBJECT.get(unit, {})
                            if electric_data:
                                inner_object.update({
                                    'fuelType': electric_data.get('fuelType'),
                                    'fuelSubType': electric_data.get('fuelSubType'),
                                    'unit': electric_data.get('unit')
                                })
                                
                        if is_reporting and is_reporting.lower() == 'yes':
                            section_three[car_type].append(inner_object)
                            
                    if section_wise_data['section_3'] and section_three:
                        section_three['is_benchmark'] = request_param.get('isBenchmark', False)
                        
                if 'section_4' in section_wise_data:
                    for s_cnt1 in range(2, len(section_wise_data['section_4'])):
                        row_data = section_wise_data['section_4'][s_cnt1]
                        sub_category = row_data[1]
                        type_ = row_data[2]
                        sub_type = row_data[3]
                        unit = row_data[4]
                        cost_per_unit = row_data[5]
                        work_place = row_data[6]
                        meter_name = row_data[7]
                        meter_number = row_data[8]
                        value = row_data[yearcnt + request_param['headerStartIndexColumnNumber']]
                        
                        car_type = re.sub(r' - ', '', re.sub(r' ', '_', type_.lower()))
                        fuel_subtype = re.sub(r' ', '_', sub_type.lower()) if sub_type else ''

                        if car_type in ['electricity', 'greenElectricity'] and work_place == 'office':
                            fuel_subtype = fuel_subtype if fuel_subtype else 'nonGrid' if car_type == 'electricity' else 'grid'

                        if unit:
                            unit = str(unit).lower()
                            
                        inner_object = {
                            'attributeCategory': category,
                            'attributeSubCategory': sub_category,
                            'fuelType': car_type,
                            'fuelSubType': fuel_subtype,
                            'scope': scope,
                            'meterName': meter_name,
                            'meterNumber': meter_number,
                            'value': value,
                            'costPerUnit': cost_per_unit,
                            'unit': unit,
                            'workPlace': work_place.lower(),
                            'isBenchmark': request_param['isBenchmark'],
                            'source': 'massUpload',
                            'modifiedDate': datetime.now()
                        }
                        section_four.setdefault(car_type, []).append(inner_object)
                    if section_wise_data['section_4'] and section_four:
                        section_four['is_benchmark'] = request_param.get('isBenchmark', False)
                        
                if 'section_5' in section_wise_data:
                    for s_cnt1 in range(2, len(section_wise_data['section_5'])):
                        row_data = section_wise_data['section_5'][s_cnt1]
                        sub_category = row_data[1]
                        type_ = row_data[2]
                        sub_type = row_data[3]
                        unit = row_data[4]
                        cost_per_unit = row_data[5]
                        work_place = row_data[6]
                        custom_type = row_data[7]
                        carbon_factor = row_data[8]
                        value = row_data[yearcnt + request_param['headerStartIndexColumnNumber']]
                        
                        car_type = re.sub(r' - ', '', to_camel_case(type_))

                        if car_type not in section_four:
                            section_four[car_type] = []

                        if unit:
                            unit = unit.lower()

                        fuel_subtype = to_camel_case(sub_type) if sub_type else ''

                        if car_type == 'fugitiveEmissions':
                            fuel_subtype = fuel_subtype.lower()
                            
                        default_subtypes = {
                            'fuelsGaseous': 'lpg',
                            'fuelsLiquid': 'dieselBioFuelBlend',
                            'bioFuels': 'bioDieselMe',
                            'biomass': 'woodPellets',
                            'biogas': 'biogas',
                            'fugitiveEmissions': 'r410a'
                        }
                        if car_type in default_subtypes and not fuel_subtype:
                            fuel_subtype = default_subtypes[car_type]
                            
                        inner_object = {
                            'attributeCategory': category,
                            'attributeSubCategory': to_camel_case(sub_category),
                            'fuelType': car_type,
                            'fuelSubType': fuel_subtype,
                            'scope': scope,
                            'value': value,
                            'costPerUnit': cost_per_unit,
                            'unit': unit,
                            'workPlace': work_place.lower(),
                            'isBenchmark': request_param.get('isBenchmark', False),
                            'source': 'massUpload',
                            'modifiedDate': datetime.now()
                        }
                        if fuel_subtype == 'custom':
                            inner_object['carbonFactor'] = carbon_factor
                            inner_object['customType'] = custom_type
                        
                        section_four[car_type].append(inner_object)
                    if len(section_wise_data['section_5']) > 0 and section_four:
                        section_four['is_benchmark'] = request_param.get('isBenchmark', False)
                        
                if 'section_6' in section_wise_data:
                    for s_cnt1 in range(2, len(section_wise_data['section_6'])):
                        row_data = section_wise_data['section_6'][s_cnt1]
                        section = row_data[0]
                        sub_category = row_data[1]
                        type_ = row_data[2]
                        sub_type = row_data[3]
                        unit = row_data[4]
                        cost_per_unit = row_data[5]
                        co2_per_unit = row_data[6]
                        value = row_data[yearcnt + request_param['headerStartIndexColumnNumber']]
                        
                        if unit is not None and unit not in [None, '']:
                            unit = str(unit).lower()

                        if (type_ and type_.lower() == 'custom') or (section and section.lower() == 'custom'):
                            if 'custom' not in section_five:
                                section_five['custom'] = []
                            
                            custom_data = {
                                "attributeCategory": request_param['category'],
                                "attributeSubCategory": sub_category,
                                "category": sub_category,  # subType replaced with subCategory
                                "subType": sub_type,
                                "scope": scope,
                                "unit": unit,
                                "value": value,
                                "co2PerUnit": co2_per_unit,
                                "costPerUnit": cost_per_unit
                            }
                            if section and section.lower() == 'custom' and type_:
                                custom_data['type'] = type_
                
                            section_five['custom'].append(custom_data)
            
                        elif sub_category and sub_category.lower() == 'direct carbon':
                            direct_carbon = value
                            
                        else:
                            sub_category_field = to_camel_case(sub_category).replace(' - ', '')
                            type_field = to_camel_case(type_).replace(' - ', '')

                            if sub_category_field not in section_five:
                                section_five[sub_category_field] = []
                            if type_field not in section_five:
                                section_five[type_field] = []

                            inner_object = {
                                "attributeCategory": request_param['category'],
                                "attributeSubCategory": sub_category_field,
                                "type": type_field,
                                "subType": to_camel_case(sub_type) if sub_type else None,
                                "scope": scope,
                                "unit": unit,
                                "costPerUnit": cost_per_unit,
                                "value": value,
                                "isBenchmark": request_param['isBenchmark'],
                                "source": 'massUpload',
                                "modifiedDate": datetime.now()
                            }
                            
                            section_five[type_field].append(inner_object)
                    if len(section_wise_data['section_6']) > 0 and len(section_five) > 0:
                        section_five['is_benchmark'] = request_param.get('isBenchmark', False)
            

        final_obj['uploadType'] = 'category'
        if section_one:
            final_obj['companyInfo'] = section_one
        if section_two:
            final_obj['employeeTravel'] = section_two
        if section_three:
            final_obj['businessTravel'] = section_three
        if section_four:
            final_obj['energyFuels'] = section_four
        if section_four:
            final_obj['detailedCarbon'] = section_five
            
        # print(direct_carbon,"111111111")
        if direct_carbon is not None:
            if 'detailedCarbon' not in final_obj:
                final_obj['detailedCarbon'] = {'is_benchmark': request_param['isBenchmark']}
                final_obj['carbonDirectCarbon'] = direct_carbon
            else:
                final_obj['carbonDirectCarbon'] = direct_carbon
                
        # Check if 'type' exists in custom project and insert flag
        if final_obj.get('detailedCarbon', {}).get('custom', []):
            for item in final_obj['detailedCarbon']['custom']:
                if 'type' not in item:
                    final_obj['customTypeExists'] = False
                    break
            
    return final_obj
    
def lower_case(value: str) -> str:
    """Convert a string to lowercase and strip whitespace."""
    return re.sub(r'\s+', ' ', value.strip().lower())

def reverseObject(obj):
    return {v: k for k, v in obj.items()}

def get_attribute_subcategories(config: Dict[str, Any]) -> List[str]:
    """
    Recursively traverses the config to extract attribute subcategories.
    """
    subcategories = []

    def traverse(obj):
        for key, value in obj.items():
            if key == "attributeSubCategory" and isinstance(value, dict):
                for sub_key, sub_value in value.items():
                    if sub_key != "custom" and sub_value != "userdefine":
                        subcategories.append(sub_value.get("label", ""))
            elif isinstance(value, dict):
                traverse(value)

    traverse(config)
    return subcategories

def excelDateToFormatedDate(excel_date: float) -> str:
    """Convert Excel date to a JS-like date string."""
    start_date = datetime(1899, 12, 30)
    result_date = start_date + timedelta(days=excel_date)
    return result_date.strftime('%b-%Y')

def validateDate(date_str: str) -> bool:
    """Validate date in the format Month-Year (e.g., Jan-2019)."""
    return bool(re.match(r'^[a-zA-Z]{3}-\d{4}$', date_str))

def get_month(month_str):
    try:
        return datetime.strptime(month_str, '%b').month - 1
    except ValueError:
        return None
    
def to_camel_case(text):
    words = text.split(' ')
    return words[0].lower() + ''.join(word.capitalize() for word in words[1:])